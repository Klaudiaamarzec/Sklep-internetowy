import time
from generator import LCG
from shop import Shop
from product import Product
# from client import Client
from openpyxl import load_workbook


def printProducts(products):
    # Petla po liscie products do wyswietlania
    for product in products:
        print(product)


def saveProducts(products, column, month):
    # zapis zebranych danych z pojedynczego dnia do pliku
    wb = load_workbook(r"niduc_projekt.xlsx")
    sheet = wb[month]
    row = 7

    for product in products:
        # Uzueplnienie kolumny w w Excelu
        sheet.cell(row, column).value = product.value1
        column = column + 1  # Przejscie do d
        # Uzupelnienie kolumny d
        sheet.cell(row, column).value = product.value2
        column = column + 1  # Przejscie do k
        # Uzupelnienie kolumny k
        sheet.cell(row, column).value = product.value3
        column = column - 2  # Powrot do pierwszej kolumny w
        row = row + 1  # Kolejny produkt

    wb.save("niduc_projekt.xlsx")


def saveErrors(error, column, month, row):
    wb = load_workbook(r"niduc_projekt.xlsx")
    sheet = wb[month]
    sheet.cell(row, column).value = error
    wb.save("niduc_projekt.xlsx")


def oneDayData(products, shop, i, month, clients_per_day, max_clients_per_product):
    # Kod generuje symulowane dane dla trzech zdarzeń: wyswietlania, dodania do koszyka i zakupu
    # Zdarzenia, ktore moga wystapic:

    # 0 - Nie wybrano żadnej opcji
    # 1 - Wyświetlenie
    # 2 - Wyświetlenie i dodanie do koszyka
    # 3 - Wyświetlenie, dodanie do koszyka i zakup

    generator = LCG()
    start_time = time.time()
    duration = 100                  # czas trwania petli w sekundach - czas trwania dnia (ustawic na 1000)
    
    row = 1010
    sum_clients = 0
    max_clients = 0
    
    # Petla powtarzajaca sie tyle czasu ile zadano
    while (time.time() - start_time) < duration:

        if shop.breakTime == 0:
            # Wyswietlenie produktu
            # Losowanie indeksu w tablicy produktow - zakres wiekszy  o 100 niz ilosc elementow w tablicy po to aby nie zawsze udalo sie wylosowac produkt
            los = generator.rand(0, len(products) + 100000)
            if los < len(products):
                products[los].value1 = products[los].value1 + 1
                shop.clients = shop.clients + 1

            # Wyswietlenie i dodanie produktu do koszyka
            los = generator.rand(0, len(products) + 1000000)
            if (los < len(products) and products[los].value2 < products[los].value1):
                products[los].value2 = products[los].value2 + 1

            # Wyswietlenie i zakup
            los = generator.rand(0, len(products) + 1000000)
            if (los < len(products) and products[los].value3 < products[los].value2):
                products[los].value3 = products[los].value3 + 1
                products[los].amount = products[los].amount - 1 #zmniejszenie ilosci produktu o 1

            # Wylosowanie błędu serwera
            # mozliwych błędów jest 15
            errorparameter = 0.3  # parametr błędu [zmierzyc jeszcze na parametrach np niski 0.1, sredni 0.5 , wysoki- 1]
            errorrange = 100000 * (1 - (
            errorparameter / 10))  # Zakres 10^9 dac - za duzo bledow wyskakuje, ale na razie tak zostawic do testowania
            errorrangefrom = 100000 * (errorparameter / 1000)
            error = generator.rand((errorrangefrom), errorrange)
            # Wylosowanie błędu serwera
            # w naszej symulacji mozliwych jest 15 bledow
            # 6 z nich jest opisane osobnymi funkcjami i prezentuja konkretne zdarzenia w symulacji
            # pozostale 9 sa losowane za pomoca generatora liczb pseudolosowych ze wzgledu na ich niezaleznosc w stosunku do uproszczen zastosowanych w symulacji
             
            if(error304(shop) == True):
                shop.error = 304
                saveErrors(str(304), i, month, row)
                row = row + 1

            if(error400(products) == True):
                shop.error = 400
                saveErrors(str(400), i, month, row)
                row = row + 1

            if(error404(products, generator) == True):
                shop.error = 404
                saveErrors(str(404), i, month, row)
                row = row + 1

            if(error408(products, month, clients_per_day) == True):
                shop.error = 408
                saveErrors(str(408), i, month, row)
                row = row + 1

            if(error410(products, generator) == True):
                shop.error = 410
                saveErrors(str(410), i, month, row)
                row = row + 1

            if(error429(products, month, max_clients_per_product) == True):
                shop.error = 429
                saveErrors(str(429), i, month, row)
                row = row + 1
            
            error = generator.rand(0, 10000000)  # Zakres 10^9 dac
            if error == 401 or error == 403 or error == 405 or error == 499 or error == 500 or error == 501 or error == 502 or error == 503 or error == 504:
                shop.error = error
                saveErrors(error, i, month, row)
                row = row + 1

            # Sprawdzanie warunkow przez caly czas dzialania sklepu
            checkConditions(products, shop)
        else:
            time.sleep(shop.breakTime)
            shop.breakTime = 0

    # Zapisanie danych z jednego dnia do pliku
    saveProducts(products, i, month)
    
    # Zliczanie ruchu na stronie w ciagu jednego dnia oraz najwiekszego ruchu w obrebie jednej podstrony
    for product in products:
        sum_clients = sum_clients + product.value1
        if(product.value1 > max_clients):
            max_clients = product.value1
    clients_per_day.append(sum_clients)
    max_clients_per_product.append(max_clients)
    sum_clients = 0
    max_clients = 0

    # Zerowanie danych na koniec dnia
    for product in products:
        product.clear()
    shop.clients = 0

# funkcje error zwracaja true jesli dany blad wystapił oraz false gdy nie wystąpił
def error304(shop):
    # 304 – Not modified
    # zasoby przechowywane w pamięci podręcznej przeglądarki nie uległy zmianie 
    # tego typu problemem może być błędna konfiguracja serwera
    # w naszej symulacji error 304 wywoła brak aktualnie dzialajacego serwera
    if(shop.runServers > 0):
        return False
    else:
        return True

def error400(products):
    # 400 – Bad request
    # niepoprawne żądanie, które zostało odrzucone przez serwer
    # w naszej symulacji jest to zdarzenie polegajace na dodaniu do koszyka produktu ktorego ilosc jest rowna 0
    for product in products:
        if(product.amount <=10 and product.value2 > 0):
            return True
    return False

def error404(products, generator):
    # 404 – Not found
    # żądany zasób nie istnieje, a dodatkowo serwer nie znajduje informacji, czy kiedykolwiek był dostępny
    # w naszej symulacji jest losowany indeks produktu, a nastepnie sprawdzane jest, czy taki indeks znajduje sie w bazie

    '''idx = generator.rand(5249, 9996858) # losowanie z zakresu indeksow
    x = 0 # licznik
    for product in products:
        if(product.index == idx):
            return False
    return True'''

    idx = generator.rand(0, len(products) + 100)
    if (idx < len(products)):
        return False
    else:
        return True
    
def error408(products, month, clients_per_day):
    # 408 – Request timeout
    # serwer przekroczył limit czasu oczekiwania na pełne żądanie przeglądarki - np. poprzez przeciążenie sieci
    # w naszej symulacji ruch czterokrotnie wiekszy na stronie niz srednia ostatnich 10 dni spowoduje przeciazenie sieci
    
    # dla pierwszych 10 dni symulacji przyjmujemy sredni ruch rowny 500
    if(month == "listopad"):
        avg_clients = 500
    # obliczenie sredniego ruchu w ciagu ostatnich 10 dni
    else:
        avg_clients = 0
        for i in range(-1, -11, -1):
            avg_clients = avg_clients + clients_per_day[i]
        avg_clients = avg_clients/10
    
    # obliczenie sumy wyswietlen w ciagu konkretnego dnia
    sum = 0
    for product in products:
        sum = sum + product.value1
    
    # warunek wystapienia bledu 408
    if(sum > 4*avg_clients):
        return True
    else:
        return False

def error410(products, generator):
    # 410 – Gone
    # żądany zasób nie istnieje, jednak serwer posiada o tym informacje (pokrewieństwo błędu 404)
    # w naszej symulacji jest to zdarzenie polegajace na wylosowaniu indeksu produktu do usuniecia
    # jesli indeks znajduje sie w bazie danych to produkt zostaje usuniety, a proba jego wyswietlenia powoduje wywolanie bledu
    idx = generator.rand(5249, 9996858) # losowanie z zakresu indeksow
    for product in products:
        if(product.index == idx and product.value1>0):
            return True
    return False

def error429(products, month, max_clients_per_product):
    # 429 – Too many requests 
    # użytkownik wysłał zbyt wiele żądań w określonym czasie
    # w naszej symulacji ruch dwurkotnie wiekszy w obrebie jednego produkty niz sredni maksymalny dzienny ruch w ostatnich 10 dniach 

    # dla pierwszych 10 dni symulacji przyjmujemy sredni maksymalny ruch w obrebie jednej podstrony rowny 15
    if(month == "listopad"):
        avg_clients_per_product = 15
    # obliczanie sredniego maksymalnego dziennego ruchu w obrebie jednej podstrony w przeciagu 10 dni
    else:
        avg_clients_per_product = 0
        for i in range(-1, -11, -1):
            avg_clients_per_product = avg_clients_per_product + max_clients_per_product[i]
        avg_clients_per_product = avg_clients_per_product/10
    
    #obliczenie maksymalnego ruchu na jednej podstronie w ciagu dnia
    max_clients = 0
    for product in products:
        if(product.value1 > max_clients):
            max_clients = product.value1
    
    #warunek wystapienia error 429
    if(max_clients > 2*avg_clients_per_product):
        return True
    else:
        return False 
    
def createData(products, shop):
    clients_per_day = [] #lista przechowujaca ilosc klientow w ciagu dni 
    max_clients_per_product = [] #lista przechowujaca maksymalna ilosc klientow danego produktu (podstrony) w ciagu dnia
    
    for i in range(6, 34, 3):  # od kolumny 6 do 33 w excelu - dni od 21 do 30 w listopadzie
        oneDayData(products, shop, i, "listopad", clients_per_day, max_clients_per_product)
    for j in range (6,97,3):
        oneDayData(products, shop, j, "grudzien", clients_per_day, max_clients_per_product)

    for k in range (6,97,3):
        oneDayData(products, shop, k, "styczen", clients_per_day, max_clients_per_product)

    for l in range (6,67,3):
        oneDayData(products, shop, l, "luty", clients_per_day, max_clients_per_product)


def read_promo_file(filename='promo.txt'):
    # pobieranie danych promocyjnych z pliku txt
    data = []
    with open(filename, 'r') as f:
        for line in f:
            day, promo, chain = line.strip().split(';')
            promo = float(promo)
            chain = chain.strip()
            data.append((int(day), promo, chain))
    return data


def checkConditions(products, shop):
    # Sprawdzanie po kolei kazdego warunku

    # Sprawdzenie warunkow dla klasy Product
    # Funkcja enumerate() tworzy iterator z pary (index, wartosc) - uzyta do uzyskania indeksu elementu w tablicy
    for i, prod in enumerate(products):
        prod.condition1()

    # Sprawdzenie dzialania sklepu internetowego:
    shop.shopErrors()
    shop.conditoins()
    shop.checkShop()


def main():
    num_samples = 1000  # liczba próbek, ile chcemy wygenerowac
    shop = Shop()
    # Lista produktow
    products = []
    row = 7

    # Przygotowanie arkusza danych do odczytu
    wb = load_workbook(r"niduc_projekt.xlsx")
    sheet = wb["listopad"]

    # Petla powtarzajaca sie tyle razy, ile mamy produktow i uzupelniajaca wartosci
    for i in range(num_samples):
        # Pobranie indeksu produktu z bazy
        idx = sheet.cell(row, column=2).value
        price = sheet.cell(row, column=4).value
        amount = 1000
        row = row + 1

        # Stworzenie obiektu product
        product = Product(idx, price, amount)
        products.append(product)  # Dodanie obiektu do tablicy produktów

    createData(products, shop)
    print("Szybkosc serwera: " + str(shop.speed))

'''
    data =read_promo_file()
    print("==================")
    print(data)
'''

if __name__ == "__main__":
    main()
